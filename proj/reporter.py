import datetime
import os
from pathlib import Path
from typing import Any

import openpyxl
import project_functions as pf
from GRANTA_MIScriptingToolkit import granta
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def report_recently_modified_records(
    folder_path: Path, records: list[granta.Record]
) -> Path:
    pf.logger.info(f"Creating report of {len(records)} records.")
    template_file_name = "report_template.xlsx"
    template_file_path = pf.current_directory / template_file_name
    workbook = openpyxl.load_workbook(filename=template_file_path)
    worksheet = workbook.worksheets[0]

    report_records(worksheet, records)

    file_path = save_Excel(folder_path, workbook)
    return file_path


def report_records(worksheet: Worksheet, records: list[granta.Record]) -> None:
    row = worksheet.max_row + 1
    for record in records:
        write_cell_hyperlink(
            worksheet, row, 1, record.history_identity, record.viewer_url
        )
        worksheet.cell(row, 2).value = record.short_name
        worksheet.cell(row, 3).value = record.name
        worksheet.cell(row, 4).value = record.pseudo_attributes["modifiedDate"].value
        worksheet.cell(row, 5).value = record.pseudo_attributes["lastModifier"].value
        row = row + 1
    return


def write_cell_hyperlink(
    worksheet: Worksheet, row: int, col: int, value: Any, URL: str
) -> Cell:
    cell = worksheet.cell(row=row, column=col)
    cell.value = '=HYPERLINK("{}", "{}")'.format(URL, value)
    cell.font = Font(color="0000FF")
    return cell


def save_Excel(folder_path: Path, workbook: Workbook) -> Path:
    file_name = datetime.datetime.now().strftime("%Y-%m-%d %H.%M.%S") + ".xlsx"
    file_path = folder_path / file_name
    folder_path.mkdir(exist_ok=True)
    workbook.save(filename=file_path)
    pf.logger.info(f"Saved report to '{file_path}'.")
    return file_path
